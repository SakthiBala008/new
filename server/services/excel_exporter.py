import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from typing import Dict, List, Optional
import base64
from datetime import datetime

class ExcelExporter:
    def __init__(self):
        self.workbook = None
        self.filename = None
        
    def create_portfolio_report(self, portfolio_data: Dict, 
                               holdings_data: List[Dict],
                               risk_metrics: Dict,
                               performance_data: Optional[pd.DataFrame] = None,
                               filename: Optional[str] = None) -> BytesIO:
        """
        Create comprehensive portfolio Excel report
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"portfolio_report_{timestamp}.xlsx"
        
        self.filename = filename
        self.workbook = Workbook()
        
        # Remove default worksheet
        self.workbook.remove(self.workbook.active)
        
        # Create worksheets
        self._create_summary_sheet(portfolio_data, risk_metrics)
        self._create_holdings_sheet(holdings_data)
        self._create_performance_sheet(performance_data)
        self._create_risk_analysis_sheet(risk_metrics)
        self._create_allocation_chart_sheet(holdings_data)
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        self.workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    def _create_summary_sheet(self, portfolio_data: Dict, risk_metrics: Dict):
        """Create portfolio summary sheet"""
        ws = self.workbook.create_sheet("Portfolio Summary")
        
        # Title
        ws['A1'] = "Portfolio Summary Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws.merge_cells('A1:D1')
        
        # Report date
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True)
        
        # Portfolio details
        row = 4
        headers = ["Metric", "Value", "Description", ""]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Portfolio metrics
        metrics = [
            ("Portfolio Name", portfolio_data.get('name', 'N/A'), "Name of the portfolio"),
            ("Total Value", f"₹{float(portfolio_data.get('totalValue', 0)):,.2f}", "Current market value"),
            ("Expected Return", f"{float(portfolio_data.get('expectedReturn', 0)):.2f}%", "Annualized expected return"),
            ("Risk Score", f"{float(portfolio_data.get('riskScore', 0)):.1f}/10", "Risk rating (1=Low, 10=High)"),
            ("Sharpe Ratio", f"{float(portfolio_data.get('sharpeRatio', 0)):.2f}", "Risk-adjusted return measure"),
            ("", "", ""),
            ("VaR (95%)", f"₹{risk_metrics.get('var_95', 0):,.2f}", "Value at Risk (95% confidence)"),
            ("Beta", f"{risk_metrics.get('beta', 0):.2f}", "Market sensitivity"),
            ("Maximum Drawdown", f"{risk_metrics.get('max_drawdown', 0):.2%}", "Largest historical decline"),
            ("Volatility", f"{risk_metrics.get('annual_volatility', 0):.2%}", "Annualized volatility"),
        ]
        
        for i, (metric, value, description) in enumerate(metrics, 5):
            ws.cell(row=i, column=1, value=metric).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)
            ws.cell(row=i, column=3, value=description)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 5
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=4, max_row=14, min_col=1, max_col=3):
            for cell in row:
                cell.border = thin_border
    
    def _create_holdings_sheet(self, holdings_data: List[Dict]):
        """Create holdings detail sheet"""
        ws = self.workbook.create_sheet("Holdings Detail")
        
        # Title
        ws['A1'] = "Portfolio Holdings"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws.merge_cells('A1:H1')
        
        # Headers
        headers = ["Symbol", "Asset Type", "Quantity", "Avg Price", "Current Value", 
                  "Allocation %", "P&L", "P&L %"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Holdings data
        total_value = sum(float(holding.get('currentValue', 0)) for holding in holdings_data)
        
        for row, holding in enumerate(holdings_data, 4):
            current_value = float(holding.get('currentValue', 0))
            avg_price = float(holding.get('avgPrice', 0))
            quantity = float(holding.get('quantity', 0))
            
            # Calculate P&L
            cost_value = avg_price * quantity
            pnl = current_value - cost_value
            pnl_percent = (pnl / cost_value * 100) if cost_value > 0 else 0
            
            ws.cell(row=row, column=1, value=holding.get('symbol', ''))
            ws.cell(row=row, column=2, value=holding.get('assetType', ''))
            ws.cell(row=row, column=3, value=quantity)
            ws.cell(row=row, column=4, value=f"₹{avg_price:,.2f}")
            ws.cell(row=row, column=5, value=f"₹{current_value:,.2f}")
            ws.cell(row=row, column=6, value=f"{float(holding.get('allocation', 0)):.2f}%")
            ws.cell(row=row, column=7, value=f"₹{pnl:,.2f}")
            ws.cell(row=row, column=8, value=f"{pnl_percent:.2f}%")
            
            # Color code P&L
            pnl_cell = ws.cell(row=row, column=7)
            pnl_percent_cell = ws.cell(row=row, column=8)
            
            if pnl >= 0:
                pnl_cell.font = Font(color="00AA00")
                pnl_percent_cell.font = Font(color="00AA00")
            else:
                pnl_cell.font = Font(color="FF0000")
                pnl_percent_cell.font = Font(color="FF0000")
        
        # Total row
        total_row = len(holdings_data) + 5
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=f"₹{total_value:,.2f}").font = Font(bold=True)
        ws.cell(row=total_row, column=6, value="100.00%").font = Font(bold=True)
        
        # Adjust column widths
        for col in range(1, 9):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=3, max_row=total_row, min_col=1, max_col=8):
            for cell in row:
                cell.border = thin_border
    
    def _create_performance_sheet(self, performance_data: Optional[pd.DataFrame]):
        """Create performance analysis sheet"""
        ws = self.workbook.create_sheet("Performance Analysis")
        
        # Title
        ws['A1'] = "Performance Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws.merge_cells('A1:D1')
        
        if performance_data is not None:
            # Add performance data
            for r in dataframe_to_rows(performance_data, index=True, header=True):
                ws.append(r)
            
            # Create performance chart
            chart = LineChart()
            chart.title = "Portfolio Performance"
            chart.style = 13
            chart.x_axis.title = "Date"
            chart.y_axis.title = "Portfolio Value"
            
            # Assuming first column is date and second is portfolio value
            data = Reference(ws, min_col=2, min_row=3, max_col=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            
            ws.add_chart(chart, "F3")
        else:
            ws['A3'] = "No performance data available"
    
    def _create_risk_analysis_sheet(self, risk_metrics: Dict):
        """Create risk analysis sheet"""
        ws = self.workbook.create_sheet("Risk Analysis")
        
        # Title
        ws['A1'] = "Risk Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws.merge_cells('A1:C1')
        
        # Risk metrics
        row = 3
        headers = ["Risk Metric", "Value", "Interpretation"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        risk_interpretations = {
            'var_95': 'Maximum expected loss (95% confidence)',
            'var_99': 'Maximum expected loss (99% confidence)',
            'beta': 'Market sensitivity (>1 = more volatile)',
            'alpha': 'Excess return vs market',
            'sharpe_ratio': 'Risk-adjusted return (>1 = good)',
            'max_drawdown': 'Largest historical decline',
            'annual_volatility': 'Yearly price fluctuation',
            'tracking_error': 'Deviation from benchmark'
        }
        
        for i, (metric, value) in enumerate(risk_metrics.items(), 4):
            if metric in risk_interpretations:
                ws.cell(row=i, column=1, value=metric.replace('_', ' ').title())
                
                if isinstance(value, (int, float)):
                    if 'percent' in metric or 'ratio' in metric:
                        ws.cell(row=i, column=2, value=f"{value:.2%}")
                    elif 'var' in metric or 'value' in metric:
                        ws.cell(row=i, column=2, value=f"₹{value:,.2f}")
                    else:
                        ws.cell(row=i, column=2, value=f"{value:.4f}")
                else:
                    ws.cell(row=i, column=2, value=str(value))
                
                ws.cell(row=i, column=3, value=risk_interpretations[metric])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.border = thin_border
    
    def _create_allocation_chart_sheet(self, holdings_data: List[Dict]):
        """Create asset allocation chart sheet"""
        ws = self.workbook.create_sheet("Asset Allocation")
        
        # Title
        ws['A1'] = "Asset Allocation"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Prepare allocation data
        allocation_data = {}
        for holding in holdings_data:
            asset_type = holding.get('assetType', 'Unknown')
            allocation = float(holding.get('allocation', 0))
            
            if asset_type in allocation_data:
                allocation_data[asset_type] += allocation
            else:
                allocation_data[asset_type] = allocation
        
        # Add data to worksheet
        ws['A3'] = "Asset Type"
        ws['B3'] = "Allocation %"
        
        for i, (asset_type, allocation) in enumerate(allocation_data.items(), 4):
            ws.cell(row=i, column=1, value=asset_type)
            ws.cell(row=i, column=2, value=allocation)
        
        # Create pie chart
        pie_chart = PieChart()
        pie_chart.title = "Asset Allocation"
        
        labels = Reference(ws, min_col=1, min_row=4, max_row=3 + len(allocation_data))
        data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(allocation_data))
        
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(labels)
        
        ws.add_chart(pie_chart, "D3")
    
    def create_risk_report(self, risk_data: Dict, filename: Optional[str] = None) -> BytesIO:
        """Create standalone risk report"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"risk_report_{timestamp}.xlsx"
        
        self.filename = filename
        self.workbook = Workbook()
        
        # Remove default worksheet
        self.workbook.remove(self.workbook.active)
        
        # Create risk analysis sheet
        self._create_risk_analysis_sheet(risk_data)
        
        # Create Monte Carlo results sheet if available
        if 'monte_carlo_results' in risk_data:
            self._create_monte_carlo_sheet(risk_data['monte_carlo_results'])
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        self.workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    def _create_monte_carlo_sheet(self, mc_results: Dict):
        """Create Monte Carlo simulation results sheet"""
        ws = self.workbook.create_sheet("Monte Carlo Analysis")
        
        # Title
        ws['A1'] = "Monte Carlo Simulation Results"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws.merge_cells('A1:D1')
        
        # Simulation parameters
        ws['A3'] = "Simulation Parameters"
        ws['A3'].font = Font(bold=True)
        
        params = [
            ("Initial Value", f"₹{mc_results.get('initial_value', 0):,.2f}"),
            ("Time Horizon", f"{mc_results.get('time_horizon_days', 0)} days"),
            ("Number of Simulations", f"{mc_results.get('num_simulations', 0):,}"),
        ]
        
        for i, (param, value) in enumerate(params, 4):
            ws.cell(row=i, column=1, value=param).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)
        
        # Results
        ws['A8'] = "Results"
        ws['A8'].font = Font(bold=True)
        
        results = [
            ("Mean Final Value", f"₹{mc_results.get('mean_final_value', 0):,.2f}"),
            ("Median Final Value", f"₹{mc_results.get('median_final_value', 0):,.2f}"),
            ("5th Percentile", f"₹{mc_results.get('percentile_5', 0):,.2f}"),
            ("95th Percentile", f"₹{mc_results.get('percentile_95', 0):,.2f}"),
            ("Probability of Loss", f"{mc_results.get('probability_of_loss', 0):.2%}"),
            ("Expected Return", f"{mc_results.get('expected_return', 0):.2%}"),
        ]
        
        for i, (metric, value) in enumerate(results, 9):
            ws.cell(row=i, column=1, value=metric).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def export_to_base64(self, excel_buffer: BytesIO) -> str:
        """Convert Excel buffer to base64 string for web download"""
        excel_buffer.seek(0)
        excel_bytes = excel_buffer.read()
        base64_string = base64.b64encode(excel_bytes).decode('utf-8')
        return base64_string
    
    def save_to_file(self, excel_buffer: BytesIO, filepath: str):
        """Save Excel buffer to file"""
        excel_buffer.seek(0)
        with open(filepath, 'wb') as f:
            f.write(excel_buffer.read())

# Example usage
if __name__ == "__main__":
    # Sample data
    portfolio_data = {
        'name': 'Growth Portfolio',
        'totalValue': '2456780.00',
        'expectedReturn': '14.20',
        'riskScore': '6.8',
        'sharpeRatio': '1.85'
    }
    
    holdings_data = [
        {
            'symbol': 'TCS',
            'assetType': 'stock',
            'quantity': 100,
            'avgPrice': 3000,
            'currentValue': 324500,
            'allocation': 25.0
        },
        {
            'symbol': 'RIL',
            'assetType': 'stock',
            'quantity': 80,
            'avgPrice': 2200,
            'currentValue': 196540,
            'allocation': 20.0
        }
    ]
    
    risk_metrics = {
        'var_95': -45230.0,
        'beta': 1.12,
        'max_drawdown': -0.085,
        'annual_volatility': 0.153,
        'sharpe_ratio': 1.85
    }
    
    # Create exporter and generate report
    exporter = ExcelExporter()
    excel_buffer = exporter.create_portfolio_report(
        portfolio_data, holdings_data, risk_metrics
    )
    
    # Save to file
    exporter.save_to_file(excel_buffer, 'portfolio_report.xlsx')
    print("Excel report generated successfully!")
